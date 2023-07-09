import requests
from bs4 import BeautifulSoup
import time
import csv
import glob
import openpyxl
import os

def get_html(url):
    """ Get the HTML of a URL """
    response = requests.get(url)
    soup = BeautifulSoup(response.text, 'lxml') # Use lxml parser
    return soup


def create_csv(letter, stocks_list_directory):
    """ Create a CSV file for each stock price letter """
    try:
        soup = get_html('https://www.moneycontrol.com/india/stockpricequote/' + letter)
        time.sleep(2)

        links = soup.find_all('a', {'class': 'bl_12'})
        file_name = os.path.join(stocks_list_directory, letter + '.csv')
        with open(file_name, 'w', newline='') as csvfile:
            writer = csv.writer(csvfile)
            for link in links:
                writer.writerow([link.text, link['href']])
            csvfile.close()
        print("Success for ", letter)

    except Exception as e:
        print("Exception for ", letter, ": ", e)


def print_csv_columns(stocks_list_directory):
    """ Print the contents of all CSVs """
    file_path = os.path.join(stocks_list_directory, '*.csv')
    for filename in glob.glob(file_path):
        with open(filename, newline='') as csvfile:
            reader = csv.reader(csvfile, delimiter=',')
            for row in reader:
                if row[0]:
                    print(row[0], row[1])
                else:
                    continue


def scrape_table(url, stock_name, sheet_name, stocks_financials_directory, next_page = False):
    """ Scrape a table from a web page """
    soup = get_html(url)
    time.sleep(2)
    table = soup.find('table', {'class': 'mctable1'})
    if table is not None:
        rows = table.find_all('tr')

        file_path = os.path.join(stocks_financials_directory, stock_name + '.xlsx')
        if os.path.isfile(file_path) and next_page==False:
            wb = openpyxl.load_workbook(file_path)
            sheet = wb.create_sheet(sheet_name)
        elif next_page:
            wb = openpyxl.load_workbook(file_path)
            sheet = wb[sheet_name]
        else:
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = sheet_name

        if next_page:
            first_empty_col = sheet.max_column - 2
            for i, row in enumerate(rows):
                for j, el in enumerate(row):
                    if j > 2 and j < len(row) - 3:
                        cell_ref = sheet.cell(i + 1, first_empty_col + j + 1)
                        cell_ref.value = el.string
        else:
            for row in rows:
                row_list = [el.string for el in row][:-2]
                sheet.append(row_list)

        wb.save(file_path)


def scrape_quick_links(url):
    """ Scrape quick links from a web page """
    soup = get_html(url)
    time.sleep(2)
    quick_links = soup.find('div', {'class': 'quick_links clearfix'})
    if quick_links is not None:
        links = quick_links.find_all('a')

        links_dict = {}
        for link in links:
            # print("{} {}".format(link.text, link['href']))
            links_dict.update({link.text: link['href']})
        return links_dict
    else:
        return None


def get_active_href(url):
    """ Get the URL of the active page """
    soup = get_html(url)
    time.sleep(2)
    span_tag = soup.find('span', {'class': 'nextpaging'})
    if span_tag is not None:
        parent_tag = span_tag.find_previous('a')
        if parent_tag:
            href = parent_tag.get('href')
            if href and href != 'javascript:void();':
                return href
    return None
